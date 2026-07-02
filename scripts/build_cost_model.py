#!/usr/bin/env python3
"""
Builds the AWS cost model spreadsheet (artifacts/cost_model.xlsx) for the
Engineer-004 real-time analytics pipeline.

Three sheets: Assumptions (editable inputs), Capacity (derived throughput/volume),
Monthly Cost (per-component cost vs the $50K ceiling). All numbers carry a
source label: Observed / Estimated / Benchmarked / Assumed.

Run: pip install openpyxl && python build_cost_model.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = Font(name='Arial', color='0000FF')
BLACK = Font(name='Arial', color='000000')
BOLD = Font(name='Arial', bold=True)
WHITE_BOLD = Font(name='Arial', bold=True, color='FFFFFF')
HDR = PatternFill('solid', start_color='1F3864')
SUB = PatternFill('solid', start_color='D9E1F2')
YEL = PatternFill('solid', start_color='FFFF00')
CTR = Alignment(horizontal='center')
LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------- Sheet 1: Assumptions ----------
a = wb.active
a.title = 'Assumptions'
a.column_dimensions['A'].width = 42
a.column_dimensions['B'].width = 16
a.column_dimensions['C'].width = 14
a.column_dimensions['D'].width = 58

a['A1'] = 'Real-Time Analytics Pipeline — Cost Model'
a['A1'].font = Font(name='Arial', bold=True, size=14)
a['A2'] = 'Blue = input you can change | Black = formula | Yellow = key assumption'
a['A2'].font = Font(name='Arial', italic=True, size=9)

for i, h in enumerate(['Assumption', 'Value', 'Label', 'Basis / Source'], 1):
    c = a.cell(row=4, column=i, value=h)
    c.font = WHITE_BOLD; c.fill = HDR; c.alignment = CTR; c.border = BORDER

rows = [
    ('Events per day (steady state)', 50000000, 'Observed', 'Given in brief: ~50M events/day from JS SDK'),
    ('Peak spike multiplier', 10, 'Observed', 'Given in brief: 10x spikes (Black Friday, launches)'),
    ('Avg event payload size (KB)', 1.0, 'Assumed', 'Typical JSON event w/ properties+context; verify against real SDK payloads'),
    ('Number of tenants', 500, 'Observed', 'Given in brief: 500+ customers, multi-tenant'),
    ('Kinesis on-demand ingest rate ($/GB)', 0.08, 'Benchmarked', 'AWS Kinesis Data Streams on-demand list price; verify current AWS pricing'),
    ('Managed Flink cost ($/KPU-hour)', 0.11, 'Benchmarked', 'AWS Managed Service for Apache Flink list price; verify current AWS pricing'),
    ('Flink KPUs provisioned (24/7)', 8, 'Estimated', 'Sized for stateful sessionization at peak; tune after load test'),
    ('ClickHouse EC2 node type cost ($/node/mo)', 1100, 'Benchmarked', 'Approx r6i.2xlarge on-demand + EBS, us-east-1; verify current AWS pricing'),
    ('ClickHouse nodes (cluster)', 3, 'Estimated', 'Replicated 3-node cluster for HA + query throughput'),
    ('S3 storage cost ($/GB/mo)', 0.023, 'Benchmarked', 'AWS S3 Standard list price; verify current AWS pricing'),
    ('Firehose delivery cost ($/GB)', 0.029, 'Benchmarked', 'AWS Kinesis Firehose list price; verify current AWS pricing'),
    ('Redis + DynamoDB identity store ($/mo)', 1200, 'Estimated', 'ElastiCache node + DynamoDB on-demand for identity graph'),
    ('Monitoring/transfer/misc ($/mo)', 3000, 'Estimated', 'CloudWatch, data transfer, NAT, headroom buffer'),
    ('Data retention in hot store (months)', 3, 'Assumed', 'Hot analytics window; older data served from S3/warehouse'),
    ('Monthly infra budget ceiling ($)', 50000, 'Observed', 'Given in brief: $50K/month infrastructure ceiling'),
]
r = 5
for name, val, label, basis in rows:
    a.cell(row=r, column=1, value=name).font = BLACK
    vc = a.cell(row=r, column=2, value=val); vc.font = BLUE; vc.alignment = CTR
    if label == 'Assumed': vc.fill = YEL
    a.cell(row=r, column=3, value=label).font = BLACK
    a.cell(row=r, column=3).alignment = CTR
    a.cell(row=r, column=4, value=basis).font = Font(name='Arial', size=9)
    a.cell(row=r, column=4).alignment = LEFT
    for col in range(1, 5): a.cell(row=r, column=col).border = BORDER
    r += 1

ADDR = {name: f"Assumptions!B{5+i}" for i, (name, _, _, _) in enumerate(rows)}

# ---------- Sheet 2: Capacity ----------
cap = wb.create_sheet('Capacity')
cap.column_dimensions['A'].width = 40
cap.column_dimensions['B'].width = 18
cap.column_dimensions['C'].width = 12
cap.column_dimensions['D'].width = 40
cap['A1'] = 'Capacity Model'; cap['A1'].font = Font(name='Arial', bold=True, size=13)
for i, h in enumerate(['Metric', 'Value', 'Label', 'Formula basis'], 1):
    c = cap.cell(row=3, column=i, value=h); c.font = WHITE_BOLD; c.fill = HDR; c.alignment = CTR; c.border = BORDER

cap_rows = [
    ('Avg events/sec', f"={ADDR['Events per day (steady state)']}/86400", 'Observed', '50M / seconds per day'),
    ('Peak events/sec (10x)', f"=B4*{ADDR['Peak spike multiplier']}", 'Estimated', 'avg/sec x spike multiplier'),
    ('Daily ingest volume (GB)', f"={ADDR['Events per day (steady state)']}*{ADDR['Avg event payload size (KB)']}/1048576", 'Estimated', 'events/day x KB / 1,048,576'),
    ('Monthly ingest volume (GB)', "=B6*30", 'Estimated', 'daily GB x 30'),
    ('Monthly ingest volume (TB)', "=B7/1024", 'Estimated', 'monthly GB / 1024'),
    ('Hot store retained volume (TB)', f"=B7*{ADDR['Data retention in hot store (months)']}/1024", 'Estimated', 'monthly GB x retention / 1024'),
]
r = 4
for name, f, label, basis in cap_rows:
    cap.cell(row=r, column=1, value=name).font = BLACK
    cap.cell(row=r, column=2, value=f).font = BLACK; cap.cell(row=r, column=2).alignment = CTR
    cap.cell(row=r, column=3, value=label).font = BLACK; cap.cell(row=r, column=3).alignment = CTR
    cap.cell(row=r, column=4, value=basis).font = Font(name='Arial', size=9)
    for col in range(1, 5): cap.cell(row=r, column=col).border = BORDER
    r += 1

# ---------- Sheet 3: Monthly Cost ----------
mc = wb.create_sheet('Monthly Cost')
mc.column_dimensions['A'].width = 34
mc.column_dimensions['B'].width = 16
mc.column_dimensions['C'].width = 12
mc.column_dimensions['D'].width = 44
mc['A1'] = 'Monthly Cost Breakdown (AWS, us-east-1)'; mc['A1'].font = Font(name='Arial', bold=True, size=13)
for i, h in enumerate(['Component', 'Monthly $', 'Label', 'Formula basis'], 1):
    c = mc.cell(row=3, column=i, value=h); c.font = WHITE_BOLD; c.fill = HDR; c.alignment = CTR; c.border = BORDER

cost_rows = [
    ('Kinesis Data Streams (ingest)', f"=Capacity!B7*{ADDR['Kinesis on-demand ingest rate ($/GB)']}", 'Estimated', 'monthly GB x $/GB ingest'),
    ('Managed Flink (stream processing)', f"={ADDR['Managed Flink cost ($/KPU-hour)']}*{ADDR['Flink KPUs provisioned (24/7)']}*730", 'Estimated', '$/KPU-hr x KPUs x hrs/mo'),
    ('ClickHouse (hot store)', f"={ADDR['ClickHouse EC2 node type cost ($/node/mo)']}*{ADDR['ClickHouse nodes (cluster)']}", 'Estimated', '$/node x nodes'),
    ('S3 (cold store / data lake)', f"=Capacity!B9*1024*{ADDR['S3 storage cost ($/GB/mo)']}", 'Estimated', 'hot-store TB x 1024 x $/GB'),
    ('Firehose (delivery to S3)', f"=Capacity!B7*{ADDR['Firehose delivery cost ($/GB)']}", 'Estimated', 'monthly GB x $/GB'),
    ('Identity store (Redis+DynamoDB)', f"={ADDR['Redis + DynamoDB identity store ($/mo)']}", 'Estimated', 'fixed monthly'),
    ('Monitoring / transfer / misc', f"={ADDR['Monitoring/transfer/misc ($/mo)']}", 'Estimated', 'fixed monthly buffer'),
]
r = 4
for name, f, label, basis in cost_rows:
    mc.cell(row=r, column=1, value=name).font = BLACK
    mc.cell(row=r, column=2, value=f).font = BLACK; mc.cell(row=r, column=2).number_format = '$#,##0'
    mc.cell(row=r, column=3, value=label).font = BLACK; mc.cell(row=r, column=3).alignment = CTR
    mc.cell(row=r, column=4, value=basis).font = Font(name='Arial', size=9)
    for col in range(1, 5): mc.cell(row=r, column=col).border = BORDER
    r += 1

total_row = r
mc.cell(row=total_row, column=1, value='TOTAL MONTHLY').font = BOLD
tc = mc.cell(row=total_row, column=2, value=f"=SUM(B4:B{r-1})"); tc.font = BOLD; tc.number_format = '$#,##0'; tc.fill = SUB
for col in range(1, 5): mc.cell(row=total_row, column=col).border = BORDER

r += 1
mc.cell(row=r, column=1, value='Budget ceiling').font = BLACK
bc = mc.cell(row=r, column=2, value=f"={ADDR['Monthly infra budget ceiling ($)']}"); bc.font = BLACK; bc.number_format = '$#,##0'
r += 1
mc.cell(row=r, column=1, value='Headroom vs ceiling').font = BOLD
hc = mc.cell(row=r, column=2, value=f"=B{r-1}-B{total_row}"); hc.font = BOLD; hc.number_format = '$#,##0'; hc.fill = YEL
r += 1
mc.cell(row=r, column=1, value='Utilization of budget').font = BLACK
uc = mc.cell(row=r, column=2, value=f"=B{total_row}/B{total_row+1}"); uc.font = BLACK; uc.number_format = '0.0%'

r += 2
note = ('NOTE: All AWS unit prices are list-price approximations and must be verified against '
        'current AWS pricing before committing. Costs scale with real payload size — the #1 '
        'sensitivity. Reserved Instances / savings plans on ClickHouse EC2 and Flink would cut '
        'steady-state cost further.')
mc.cell(row=r, column=1, value=note).font = Font(name='Arial', italic=True, size=9)
mc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
mc.cell(row=r, column=1).alignment = LEFT

os.makedirs('artifacts', exist_ok=True)
wb.save('artifacts/cost_model.xlsx')
print("saved artifacts/cost_model.xlsx")